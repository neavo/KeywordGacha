from typing import cast

import openpyxl
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

from module.Utils.SpreadsheetTool import SpreadsheetTool


def test_get_cell_value_handles_none_and_strips_text() -> None:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    assert isinstance(sheet, Worksheet)

    cast(Cell, sheet.cell(row=1, column=1)).value = None
    cast(Cell, sheet.cell(row=1, column=2)).value = "  hello  "
    cast(Cell, sheet.cell(row=1, column=3)).value = 123

    assert SpreadsheetTool.get_cell_value(sheet, 1, 1) == ""
    assert SpreadsheetTool.get_cell_value(sheet, 1, 2) == "hello"
    assert SpreadsheetTool.get_cell_value(sheet, 1, 3) == "123"


def test_set_cell_value_handles_none_and_escapes_formula() -> None:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    assert isinstance(sheet, Worksheet)

    SpreadsheetTool.set_cell_value(sheet, 1, 1, None)
    SpreadsheetTool.set_cell_value(sheet, 1, 2, "=SUM(A1:A2)")

    assert sheet.cell(row=1, column=1).value == ""
    assert sheet.cell(row=1, column=2).value == "'=SUM(A1:A2)"


def test_set_cell_value_keeps_plain_text_and_default_font_size() -> None:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    assert isinstance(sheet, Worksheet)

    SpreadsheetTool.set_cell_value(sheet, 2, 1, "plain text")
    cell = sheet.cell(row=2, column=1)

    assert cell.value == "plain text"
    assert cell.font.size == 9


def test_set_cell_value_applies_font_and_alignment() -> None:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    assert isinstance(sheet, Worksheet)

    SpreadsheetTool.set_cell_value(sheet, 1, 1, "value", font_size=11)
    cell = sheet.cell(row=1, column=1)

    assert cell.font.size == 11
    assert cell.alignment.wrap_text is True
    assert cell.alignment.vertical == "center"
    assert cell.alignment.horizontal == "left"
