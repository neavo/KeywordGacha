import json
from pathlib import Path
from typing import cast

import openpyxl
import pytest
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

from module.QualityRule.QualityRuleIO import QualityRuleIO
from module.Utils.SpreadsheetTool import SpreadsheetTool


def test_load_rules_from_file_dispatches_by_extension(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    monkeypatch.setattr(
        QualityRuleIO,
        "load_from_json_file",
        staticmethod(lambda path: [{"src": path}]),
    )
    monkeypatch.setattr(
        QualityRuleIO,
        "load_from_xlsx_file",
        staticmethod(lambda path: [{"dst": path}]),
    )

    assert QualityRuleIO.load_rules_from_file("rules.json") == [{"src": "rules.json"}]
    assert QualityRuleIO.load_rules_from_file("rules.XLSX") == [{"dst": "rules.XLSX"}]
    assert QualityRuleIO.load_rules_from_file("") == []
    assert QualityRuleIO.load_rules_from_file("rules.txt") == []


def test_load_from_json_file_supports_list_actors_and_dict_formats(
    fs,
) -> None:
    del fs
    root_path = Path("/workspace/quality")
    root_path.mkdir(parents=True, exist_ok=True)
    list_path = root_path / "list.json"
    list_path.write_text(
        json.dumps(
            [
                {"src": " HP ", "dst": "生命值", "info": "i", "regex": 1},
                {"src": "   "},
                "bad",
            ]
        ),
        encoding="utf-8",
    )
    list_result = QualityRuleIO.load_from_json_file(str(list_path))
    assert list_result[0]["src"] == "HP"
    assert list_result[0]["regex"] is True

    actors_path = root_path / "actors.json"
    actors_path.write_text(
        json.dumps([{"id": 7, "name": "勇者", "nickname": "小勇"}]),
        encoding="utf-8",
    )
    actor_result = QualityRuleIO.load_from_json_file(str(actors_path))
    assert {item["src"] for item in actor_result} == {
        "\\n[7]",
        "\\N[7]",
        "\\nn[7]",
        "\\NN[7]",
    }

    kv_path = root_path / "kv.json"
    kv_path.write_text(json.dumps({"A": "甲", "": "skip", "B": None}), encoding="utf-8")
    kv_result = QualityRuleIO.load_from_json_file(str(kv_path))
    assert kv_result == [
        {
            "src": "A",
            "dst": "甲",
            "info": "",
            "regex": False,
            "case_sensitive": False,
        },
        {
            "src": "B",
            "dst": "",
            "info": "",
            "regex": False,
            "case_sensitive": False,
        },
    ]


def test_load_from_json_file_skips_non_string_keys(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    monkeypatch.setattr(
        "module.QualityRule.QualityRuleIO.JSONTool.load_file",
        lambda _path: {1: "one", "A": "甲", 2: "two"},
    )

    result = QualityRuleIO.load_from_json_file("/workspace/quality/kv.json")

    assert result == [
        {
            "src": "A",
            "dst": "甲",
            "info": "",
            "regex": False,
            "case_sensitive": False,
        }
    ]


def test_load_from_json_file_handles_actor_rows_with_partial_fields(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    monkeypatch.setattr(
        "module.QualityRule.QualityRuleIO.JSONTool.load_file",
        lambda _path: [
            {"id": 7, "name": "", "nickname": "小勇"},
            {"id": 8, "name": "勇者", "nickname": ""},
        ],
    )

    result = QualityRuleIO.load_from_json_file("/workspace/quality/actors.json")

    assert [item["src"] for item in result] == [
        "\\nn[7]",
        "\\NN[7]",
        "\\n[8]",
        "\\N[8]",
    ]


def test_load_from_xlsx_file_skips_header_and_parses_booleans(
    fs, monkeypatch: pytest.MonkeyPatch
) -> None:
    del fs
    book = openpyxl.Workbook()
    sheet = book.active
    assert isinstance(sheet, Worksheet)

    cast(Cell, sheet.cell(row=1, column=1)).value = "src"
    cast(Cell, sheet.cell(row=1, column=2)).value = "dst"
    cast(Cell, sheet.cell(row=2, column=1)).value = "HP"
    cast(Cell, sheet.cell(row=2, column=2)).value = "生命值"
    cast(Cell, sheet.cell(row=2, column=3)).value = "term"
    cast(Cell, sheet.cell(row=2, column=4)).value = "true"
    cast(Cell, sheet.cell(row=2, column=5)).value = "TRUE"
    cast(Cell, sheet.cell(row=3, column=1)).value = ""
    monkeypatch.setattr(
        "module.QualityRule.QualityRuleIO.openpyxl.load_workbook",
        lambda *_args, **_kwargs: book,
    )

    result = QualityRuleIO.load_from_xlsx_file("/workspace/quality/rules.xlsx")
    assert result == [
        {
            "src": "HP",
            "dst": "生命值",
            "info": "term",
            "regex": True,
            "case_sensitive": True,
        }
    ]


def test_load_from_xlsx_file_returns_empty_when_active_is_not_worksheet(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    class FakeWorkbook:
        active = "not-worksheet"

    monkeypatch.setattr(
        "module.QualityRule.QualityRuleIO.openpyxl.load_workbook",
        lambda *_args, **_kwargs: FakeWorkbook(),
    )

    assert QualityRuleIO.load_from_xlsx_file("/workspace/quality/rules.xlsx") == []


def test_load_from_xlsx_file_skips_rows_with_none_first_cell(
    fs, monkeypatch: pytest.MonkeyPatch
) -> None:
    del fs
    book = openpyxl.Workbook()
    sheet = book.active
    assert isinstance(sheet, Worksheet)

    cast(Cell, sheet.cell(row=1, column=1)).value = "src"
    cast(Cell, sheet.cell(row=1, column=2)).value = "dst"
    cast(Cell, sheet.cell(row=2, column=1)).value = "HP"
    cast(Cell, sheet.cell(row=2, column=2)).value = "生命值"
    cast(Cell, sheet.cell(row=3, column=1)).value = None
    cast(Cell, sheet.cell(row=3, column=2)).value = "应跳过"
    cast(Cell, sheet.cell(row=4, column=1)).value = "MP"
    cast(Cell, sheet.cell(row=4, column=2)).value = "魔力"

    monkeypatch.setattr(
        "module.QualityRule.QualityRuleIO.openpyxl.load_workbook",
        lambda *_args, **_kwargs: book,
    )

    result = QualityRuleIO.load_from_xlsx_file("/workspace/quality/rules.xlsx")

    assert [item["src"] for item in result] == ["HP", "MP"]


def test_load_from_xlsx_file_skips_rows_when_reader_returns_none(
    fs, monkeypatch: pytest.MonkeyPatch
) -> None:
    del fs
    book = openpyxl.Workbook()
    sheet = book.active
    assert isinstance(sheet, Worksheet)

    cast(Cell, sheet.cell(row=1, column=1)).value = "src"
    cast(Cell, sheet.cell(row=1, column=2)).value = "dst"
    cast(Cell, sheet.cell(row=2, column=1)).value = "IGNORED"
    cast(Cell, sheet.cell(row=2, column=2)).value = "忽略"

    original_get = SpreadsheetTool.get_cell_value

    def fake_get_cell_value(
        ws: Worksheet,
        row: int,
        column: int,
    ) -> str | None:
        if row == 2 and column == 1:
            return None
        return original_get(ws, row, column)

    monkeypatch.setattr(
        "module.QualityRule.QualityRuleIO.openpyxl.load_workbook",
        lambda *_args, **_kwargs: book,
    )
    monkeypatch.setattr(
        "module.QualityRule.QualityRuleIO.SpreadsheetTool.get_cell_value",
        fake_get_cell_value,
    )

    result = QualityRuleIO.load_from_xlsx_file("/workspace/quality/rules.xlsx")

    assert result == []


def test_export_rules_writes_xlsx_and_json(fs, monkeypatch: pytest.MonkeyPatch) -> None:
    del fs
    captured: dict[str, object] = {}

    def fake_save(workbook: openpyxl.Workbook, path: str) -> None:
        sheet = workbook.active
        assert isinstance(sheet, Worksheet)
        captured["path"] = path
        captured["header"] = sheet.cell(row=1, column=1).value
        captured["row2_col1"] = sheet.cell(row=2, column=1).value
        captured["row2_col2"] = sheet.cell(row=2, column=2).value
        output_file = Path(path)
        output_file.parent.mkdir(parents=True, exist_ok=True)
        output_file.write_bytes(b"xlsx")

    monkeypatch.setattr(
        "module.QualityRule.QualityRuleIO.openpyxl.Workbook.save", fake_save
    )

    output_base = Path("/workspace/quality/exported_rules")
    output_base.parent.mkdir(parents=True, exist_ok=True)
    rules = [
        {
            "src": "A",
            "dst": "甲",
            "info": "i",
            "regex": False,
            "case_sensitive": True,
        }
    ]

    QualityRuleIO.export_rules(str(output_base), rules)

    json_path = Path("/workspace/quality/exported_rules.json")
    xlsx_path = Path("/workspace/quality/exported_rules.xlsx")
    assert json_path.exists()
    assert xlsx_path.exists()

    assert json.loads(json_path.read_text(encoding="utf-8")) == rules
    assert captured["path"] == str(xlsx_path)
    assert captured["header"] == "src"
    assert captured["row2_col1"] == "A"
    assert captured["row2_col2"] == "甲"


def test_export_rules_raises_when_active_sheet_is_invalid(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    class FakeWorkbook:
        active = None

    monkeypatch.setattr(
        "module.QualityRule.QualityRuleIO.openpyxl.Workbook",
        lambda: FakeWorkbook(),
    )

    with pytest.raises(RuntimeError, match="Failed to create worksheet"):
        QualityRuleIO.export_rules("/workspace/quality/rules", [{"src": "A"}])
