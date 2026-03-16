from __future__ import annotations

import io
from pathlib import Path
from typing import cast

import openpyxl
import pytest
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

from base.Base import Base
from model.Item import Item
from module.Config import Config
from module.File.XLSX import XLSX
from tests.module.file.conftest import DummyDataManager


def build_xlsx_bytes(setup_rows: list[tuple[int, int, str | int | None]]) -> bytes:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    assert isinstance(sheet, Worksheet)
    for row, column, value in setup_rows:
        cast(Cell, sheet.cell(row=row, column=column)).value = value
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def test_read_from_stream_reads_status_for_common_rows(config: Config) -> None:
    content = build_xlsx_bytes(
        [
            (1, 1, "src1"),
            (1, 2, "dst1"),
            (2, 1, "same"),
            (2, 2, "same"),
            (3, 1, 123),
        ]
    )

    items = XLSX(config).read_from_stream(content, "sheet.xlsx")

    assert len(items) == 3
    assert items[0].get_status() == Base.ProjectStatus.PROCESSED_IN_PAST
    assert items[1].get_status() == Base.ProjectStatus.NONE
    assert items[2].get_src() == "123"
    assert items[2].get_dst() == ""


def test_read_from_stream_skips_rows_with_missing_source(config: Config) -> None:
    content = build_xlsx_bytes(
        [
            (1, 1, "src1"),
            (1, 2, "dst1"),
            (3, 1, "src3"),
        ]
    )

    items = XLSX(config).read_from_stream(content, "sheet.xlsx")

    assert len(items) == 2
    assert [item.get_src() for item in items] == ["src1", "src3"]


def test_read_from_stream_marks_empty_source_as_excluded(
    config: Config,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    assert isinstance(sheet, Worksheet)
    cast(Cell, sheet.cell(row=1, column=1)).value = ""
    cast(Cell, sheet.cell(row=1, column=2)).value = "x"

    monkeypatch.setattr("module.File.XLSX.openpyxl.load_workbook", lambda *_: workbook)

    items = XLSX(config).read_from_stream(b"bytes", "sheet.xlsx")

    assert len(items) == 1
    assert items[0].get_status() == Base.ProjectStatus.EXCLUDED


def test_read_from_stream_skips_wolf_sheet(config: Config) -> None:
    content = build_xlsx_bytes(
        [
            (1, 1, "code"),
            (1, 2, "flag"),
            (1, 3, "type"),
            (1, 4, "info"),
            (2, 1, "src"),
            (2, 2, "dst"),
        ]
    )

    assert XLSX(config).read_from_stream(content, "wolf.xlsx") == []


def test_is_wold_xlsx_detects_required_headers(config: Config) -> None:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    assert isinstance(sheet, Worksheet)
    cast(Cell, sheet.cell(row=1, column=1)).value = "Code"
    cast(Cell, sheet.cell(row=1, column=2)).value = "Flag"
    cast(Cell, sheet.cell(row=1, column=3)).value = "Type"
    cast(Cell, sheet.cell(row=1, column=4)).value = "Info"

    assert XLSX(config).is_wold_xlsx(sheet) is True


@pytest.mark.parametrize(
    "col,value",
    [
        (1, "wrong"),
        (2, "wrong"),
        (3, "wrong"),
        (4, "wrong"),
    ],
)
def test_is_wold_xlsx_returns_false_when_header_missing(
    col: int,
    value: str,
    config: Config,
) -> None:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    assert isinstance(sheet, Worksheet)
    cast(Cell, sheet.cell(row=1, column=1)).value = "Code"
    cast(Cell, sheet.cell(row=1, column=2)).value = "Flag"
    cast(Cell, sheet.cell(row=1, column=3)).value = "Type"
    cast(Cell, sheet.cell(row=1, column=4)).value = "Info"
    cast(Cell, sheet.cell(row=1, column=col)).value = value

    assert XLSX(config).is_wold_xlsx(sheet) is False


def test_write_to_path_writes_sorted_rows_to_excel(
    config: Config,
    dummy_data_manager: DummyDataManager,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    captured: dict[str, object] = {}

    def fake_save(workbook: openpyxl.Workbook, path: str) -> None:
        sheet = workbook.active
        assert isinstance(sheet, Worksheet)
        captured["path"] = path
        captured["row1_col1"] = sheet.cell(row=1, column=1).value
        captured["row1_col2"] = sheet.cell(row=1, column=2).value
        captured["row2_col1"] = sheet.cell(row=2, column=1).value
        captured["row2_col2"] = sheet.cell(row=2, column=2).value
        captured["width_a"] = sheet.column_dimensions["A"].width
        captured["width_b"] = sheet.column_dimensions["B"].width
        output_file = Path(path)
        output_file.parent.mkdir(parents=True, exist_ok=True)
        output_file.write_bytes(b"xlsx")

    monkeypatch.setattr("module.File.XLSX.openpyxl.Workbook.save", fake_save)
    monkeypatch.setattr("module.File.XLSX.DataManager.get", lambda: dummy_data_manager)
    items = [
        Item.from_dict(
            {
                "src": "row2-src",
                "dst": "row2-dst",
                "row": 2,
                "file_type": Item.FileType.XLSX,
                "file_path": "excel/a.xlsx",
            }
        ),
        Item.from_dict(
            {
                "src": "row1-src",
                "dst": "row1-dst",
                "row": 1,
                "file_type": Item.FileType.XLSX,
                "file_path": "excel/a.xlsx",
            }
        ),
    ]

    XLSX(config).write_to_path(items)

    output_file = Path(dummy_data_manager.get_translated_path()) / "excel" / "a.xlsx"
    assert output_file.exists()
    assert str(captured["path"]).replace("\\", "/") == str(output_file).replace(
        "\\", "/"
    )
    assert captured["row1_col1"] == "row1-src"
    assert captured["row1_col2"] == "row1-dst"
    assert captured["row2_col1"] == "row2-src"
    assert captured["row2_col2"] == "row2-dst"
    assert captured["width_a"] == 64
    assert captured["width_b"] == 64


def test_write_to_path_keeps_empty_dst_as_empty_cell(
    config: Config,
    dummy_data_manager: DummyDataManager,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    captured: dict[str, object] = {}

    def fake_save(workbook: openpyxl.Workbook, path: str) -> None:
        sheet = workbook.active
        assert isinstance(sheet, Worksheet)
        captured["dst"] = sheet.cell(row=1, column=2).value
        output_file = Path(path)
        output_file.parent.mkdir(parents=True, exist_ok=True)
        output_file.write_bytes(b"xlsx")

    monkeypatch.setattr("module.File.XLSX.openpyxl.Workbook.save", fake_save)
    monkeypatch.setattr("module.File.XLSX.DataManager.get", lambda: dummy_data_manager)

    XLSX(config).write_to_path(
        [
            Item.from_dict(
                {
                    "src": "row1-src",
                    "dst": "",
                    "row": 1,
                    "file_type": Item.FileType.XLSX,
                    "file_path": "excel/empty.xlsx",
                }
            )
        ]
    )

    assert captured["dst"] == ""


def test_read_from_path_reads_files(
    fs,
    config: Config,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    input_root = Path("/fake/input")
    path = input_root / "nested" / "a.xlsx"
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_bytes(b"fake-xlsx")

    captured: dict[str, object] = {}

    def fake_read_from_stream(self: XLSX, content: bytes, rel_path: str) -> list[Item]:
        del self
        captured["content"] = content
        captured["rel_path"] = rel_path
        return [
            Item.from_dict(
                {
                    "src": "src",
                    "dst": "dst",
                    "row": 1,
                    "file_type": Item.FileType.XLSX,
                    "file_path": rel_path,
                }
            )
        ]

    monkeypatch.setattr(XLSX, "read_from_stream", fake_read_from_stream)

    items = XLSX(config).read_from_path([str(path)], str(input_root))

    assert len(items) == 1
    assert items[0].get_file_path().replace("\\", "/") == "nested/a.xlsx"
    assert captured["content"] == b"fake-xlsx"
    assert str(captured["rel_path"]).replace("\\", "/") == "nested/a.xlsx"


def test_read_from_stream_returns_empty_when_active_sheet_is_not_worksheet(
    config: Config,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    class DummyBook:
        active = object()

    monkeypatch.setattr(
        "module.File.XLSX.openpyxl.load_workbook", lambda *_: DummyBook()
    )

    assert XLSX(config).read_from_stream(b"bytes", "a.xlsx") == []


def test_read_from_stream_returns_empty_when_sheet_dimension_is_zero(
    config: Config,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    assert isinstance(sheet, Worksheet)
    cast(Cell, sheet.cell(row=1, column=1)).value = "src"

    monkeypatch.setattr("module.File.XLSX.openpyxl.load_workbook", lambda *_: workbook)
    monkeypatch.setattr(Worksheet, "max_row", property(lambda self: 0))

    assert XLSX(config).read_from_stream(b"bytes", "a.xlsx") == []


def test_write_to_path_skips_when_active_sheet_is_not_worksheet(
    config: Config,
    dummy_data_manager: DummyDataManager,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    class DummyBook:
        def __init__(self) -> None:
            self.active = object()

        def save(self, path: str) -> None:
            del path

    monkeypatch.setattr("module.File.XLSX.openpyxl.Workbook", DummyBook)
    monkeypatch.setattr("module.File.XLSX.DataManager.get", lambda: dummy_data_manager)

    XLSX(config).write_to_path(
        [
            Item.from_dict(
                {
                    "src": "row1-src",
                    "dst": "row1-dst",
                    "row": 1,
                    "file_type": Item.FileType.XLSX,
                    "file_path": "excel/not-sheet.xlsx",
                }
            )
        ]
    )

    output_file = (
        Path(dummy_data_manager.get_translated_path()) / "excel" / "not-sheet.xlsx"
    )
    assert output_file.exists() is False
