from __future__ import annotations

import io
from pathlib import Path
from typing import cast

import openpyxl
import openpyxl.styles
import pytest
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

from base.Base import Base
from model.Item import Item
from module.Config import Config
from module.File.WOLFXLSX import WOLFXLSX
from tests.module.file.conftest import DummyDataManager


def build_wolf_xlsx_bytes() -> bytes:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    assert isinstance(sheet, Worksheet)

    cast(Cell, sheet.cell(row=1, column=1)).value = "code"
    cast(Cell, sheet.cell(row=1, column=2)).value = "flag"
    cast(Cell, sheet.cell(row=1, column=3)).value = "type"
    cast(Cell, sheet.cell(row=1, column=4)).value = "info"

    white_fill = openpyxl.styles.PatternFill(
        fill_type="solid",
        fgColor=openpyxl.styles.Color(indexed=9),
    )
    blue_fill = openpyxl.styles.PatternFill(
        fill_type="solid",
        fgColor=openpyxl.styles.Color(indexed=44),
    )

    row2_src = cast(Cell, sheet.cell(row=2, column=WOLFXLSX.COL_SRC_TEXT))
    row2_src.value = "原文1"
    row2_src.fill = white_fill

    row3_src = cast(Cell, sheet.cell(row=3, column=WOLFXLSX.COL_SRC_TEXT))
    row3_src.value = "原文2"
    row3_src.fill = white_fill
    cast(Cell, sheet.cell(row=3, column=WOLFXLSX.COL_DST_TEXT)).value = "已译2"

    row4_src = cast(Cell, sheet.cell(row=4, column=WOLFXLSX.COL_SRC_TEXT))
    row4_src.value = "原文3"
    row4_src.fill = blue_fill

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def test_read_from_stream_sets_status_from_color_and_dst(config: Config) -> None:
    content = build_wolf_xlsx_bytes()

    items = WOLFXLSX(config).read_from_stream(content, "wolf.xlsx")

    assert len(items) == 3
    assert items[0].get_status() == Base.ProjectStatus.NONE
    assert items[1].get_status() == Base.ProjectStatus.PROCESSED_IN_PAST
    assert items[2].get_status() == Base.ProjectStatus.EXCLUDED


def test_get_fg_color_index_returns_minus_one_without_fill(config: Config) -> None:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    assert isinstance(sheet, Worksheet)
    cast(Cell, sheet.cell(row=2, column=6)).value = "text"

    assert WOLFXLSX(config).get_fg_color_index(sheet, 2, 6) == -1


def test_write_to_path_restores_original_workbook_when_asset_exists(
    config: Config,
    dummy_data_manager: DummyDataManager,
    fs,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    captured: dict[str, object] = {}
    fs.pause()
    try:
        original = build_wolf_xlsx_bytes()
        original_load_workbook = openpyxl.load_workbook
        restored_book = original_load_workbook(io.BytesIO(original))
    finally:
        fs.resume()

    def load_workbook_compat(path_or_stream, *args, **kwargs):
        if isinstance(path_or_stream, (str, Path)):
            del args
            del kwargs
            return restored_book
        return original_load_workbook(path_or_stream, *args, **kwargs)

    def fake_save(workbook: openpyxl.Workbook, path: str) -> None:
        sheet = workbook.active
        assert isinstance(sheet, Worksheet)
        captured["path"] = path
        captured["src"] = sheet.cell(row=2, column=WOLFXLSX.COL_SRC_TEXT).value
        captured["dst"] = sheet.cell(row=2, column=WOLFXLSX.COL_DST_TEXT).value
        output_file = Path(path)
        output_file.parent.mkdir(parents=True, exist_ok=True)
        output_file.write_bytes(b"xlsx")

    monkeypatch.setattr(
        "module.File.WOLFXLSX.openpyxl.load_workbook", load_workbook_compat
    )
    monkeypatch.setattr("module.File.WOLFXLSX.openpyxl.Workbook.save", fake_save)

    dummy_data_manager.assets["wolf/game.xlsx"] = original
    monkeypatch.setattr(
        "module.File.WOLFXLSX.DataManager.get", lambda: dummy_data_manager
    )
    items = [
        Item.from_dict(
            {
                "src": "最新原文",
                "dst": "最新译文",
                "row": 2,
                "file_type": Item.FileType.WOLFXLSX,
                "file_path": "wolf/game.xlsx",
            }
        )
    ]

    WOLFXLSX(config).write_to_path(items)

    output_file = Path(dummy_data_manager.get_translated_path()) / "wolf" / "game.xlsx"
    assert output_file.exists()
    assert str(captured["path"]).replace("\\", "/") == str(output_file).replace(
        "\\", "/"
    )
    assert captured["src"] == "最新原文"
    assert captured["dst"] == "最新译文"


def test_read_from_path_reads_files(
    fs,
    config: Config,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    input_root = Path("/fake/input")
    path = input_root / "nested" / "wolf.xlsx"
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_bytes(b"fake-wolf-xlsx")

    captured: dict[str, object] = {}

    def fake_read_from_stream(
        self: WOLFXLSX,
        content: bytes,
        rel_path: str,
    ) -> list[Item]:
        del self
        captured["content"] = content
        captured["rel_path"] = rel_path
        return [
            Item.from_dict(
                {
                    "src": "原文",
                    "dst": "译文",
                    "row": 2,
                    "file_type": Item.FileType.WOLFXLSX,
                    "file_path": rel_path,
                }
            )
        ]

    monkeypatch.setattr(WOLFXLSX, "read_from_stream", fake_read_from_stream)

    items = WOLFXLSX(config).read_from_path([str(path)], str(input_root))

    assert len(items) == 1
    assert {item.get_file_path().replace("\\", "/") for item in items} == {
        "nested/wolf.xlsx"
    }
    assert captured["content"] == b"fake-wolf-xlsx"
    assert str(captured["rel_path"]).replace("\\", "/") == "nested/wolf.xlsx"


def test_read_from_stream_returns_empty_when_active_sheet_not_worksheet(
    config: Config,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    class DummyBook:
        active = object()

    monkeypatch.setattr(
        "module.File.WOLFXLSX.openpyxl.load_workbook",
        lambda *_: DummyBook(),
    )

    assert WOLFXLSX(config).read_from_stream(b"bytes", "wolf.xlsx") == []


def test_read_from_stream_returns_empty_when_not_wolf_sheet(config: Config) -> None:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    assert isinstance(sheet, Worksheet)
    cast(Cell, sheet.cell(row=1, column=1)).value = "not-code"
    buffer = io.BytesIO()
    workbook.save(buffer)

    assert WOLFXLSX(config).read_from_stream(buffer.getvalue(), "wolf.xlsx") == []


def test_read_from_stream_skips_row_when_source_is_none(config: Config) -> None:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    assert isinstance(sheet, Worksheet)

    cast(Cell, sheet.cell(row=1, column=1)).value = "code"
    cast(Cell, sheet.cell(row=1, column=2)).value = "flag"
    cast(Cell, sheet.cell(row=1, column=3)).value = "type"
    cast(Cell, sheet.cell(row=1, column=4)).value = "info"

    white_fill = openpyxl.styles.PatternFill(
        fill_type="solid",
        fgColor=openpyxl.styles.Color(indexed=9),
    )

    # row=2 留空，触发 src_val is None 分支。
    row3_src = cast(Cell, sheet.cell(row=3, column=WOLFXLSX.COL_SRC_TEXT))
    row3_src.value = "原文"
    row3_src.fill = white_fill

    buffer = io.BytesIO()
    workbook.save(buffer)

    items = WOLFXLSX(config).read_from_stream(buffer.getvalue(), "wolf.xlsx")

    assert len(items) == 1
    assert items[0].get_src() == "原文"


def test_get_fg_color_index_returns_minus_one_for_non_indexed_color(
    config: Config,
) -> None:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    assert isinstance(sheet, Worksheet)

    cell = cast(Cell, sheet.cell(row=2, column=6))
    cell.value = "text"
    cell.fill = openpyxl.styles.PatternFill(
        fill_type="solid",
        fgColor=openpyxl.styles.Color(rgb="FFFFFF"),
    )

    assert WOLFXLSX(config).get_fg_color_index(sheet, 2, 6) == -1


def test_write_to_path_falls_back_to_new_workbook_when_asset_missing(
    config: Config,
    dummy_data_manager: DummyDataManager,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    captured: dict[str, object] = {}

    def fake_save(workbook: openpyxl.Workbook, path: str) -> None:
        sheet = workbook.active
        assert isinstance(sheet, Worksheet)
        captured["path"] = path
        captured["width_a"] = sheet.column_dimensions["A"].width
        captured["width_b"] = sheet.column_dimensions["B"].width
        captured["src"] = sheet.cell(row=2, column=WOLFXLSX.COL_SRC_TEXT).value
        captured["dst"] = sheet.cell(row=2, column=WOLFXLSX.COL_DST_TEXT).value
        output_file = Path(path)
        output_file.parent.mkdir(parents=True, exist_ok=True)
        output_file.write_bytes(b"xlsx")

    monkeypatch.setattr("module.File.WOLFXLSX.openpyxl.Workbook.save", fake_save)
    monkeypatch.setattr(
        "module.File.WOLFXLSX.DataManager.get", lambda: dummy_data_manager
    )

    items = [
        Item.from_dict(
            {
                "src": "原文",
                "dst": "译文",
                "row": 2,
                "file_type": Item.FileType.WOLFXLSX,
                "file_path": "wolf/game.xlsx",
            }
        )
    ]

    WOLFXLSX(config).write_to_path(items)

    output_file = Path(dummy_data_manager.get_translated_path()) / "wolf" / "game.xlsx"
    assert output_file.exists()
    assert str(captured["path"]).replace("\\", "/") == str(output_file).replace(
        "\\", "/"
    )
    assert captured["width_a"] == 64
    assert captured["width_b"] == 64
    assert captured["src"] == "原文"
    assert captured["dst"] == "译文"


def test_write_to_path_keeps_empty_dst_as_empty_cell(
    config: Config,
    dummy_data_manager: DummyDataManager,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    captured: dict[str, object] = {}

    def fake_save(workbook: openpyxl.Workbook, path: str) -> None:
        sheet = workbook.active
        assert isinstance(sheet, Worksheet)
        captured["dst"] = sheet.cell(row=2, column=WOLFXLSX.COL_DST_TEXT).value
        output_file = Path(path)
        output_file.parent.mkdir(parents=True, exist_ok=True)
        output_file.write_bytes(b"xlsx")

    monkeypatch.setattr("module.File.WOLFXLSX.openpyxl.Workbook.save", fake_save)
    monkeypatch.setattr(
        "module.File.WOLFXLSX.DataManager.get", lambda: dummy_data_manager
    )

    WOLFXLSX(config).write_to_path(
        [
            Item.from_dict(
                {
                    "src": "原文",
                    "dst": "",
                    "row": 2,
                    "file_type": Item.FileType.WOLFXLSX,
                    "file_path": "wolf/empty.xlsx",
                }
            )
        ]
    )

    assert captured["dst"] == ""


def test_read_from_stream_returns_empty_when_sheet_dimension_is_zero(
    config: Config,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    assert isinstance(sheet, Worksheet)
    cast(Cell, sheet.cell(row=1, column=1)).value = "code"
    cast(Cell, sheet.cell(row=1, column=2)).value = "flag"
    cast(Cell, sheet.cell(row=1, column=3)).value = "type"
    cast(Cell, sheet.cell(row=1, column=4)).value = "info"

    monkeypatch.setattr(
        "module.File.WOLFXLSX.openpyxl.load_workbook", lambda *_: workbook
    )
    monkeypatch.setattr(Worksheet, "max_row", property(lambda self: 0))

    assert WOLFXLSX(config).read_from_stream(b"bytes", "wolf.xlsx") == []


def test_write_to_path_skips_when_new_workbook_active_is_not_worksheet(
    config: Config,
    dummy_data_manager: DummyDataManager,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    class DummyBook:
        def __init__(self) -> None:
            self.active = object()

    monkeypatch.setattr("module.File.WOLFXLSX.openpyxl.Workbook", DummyBook)
    monkeypatch.setattr(
        "module.File.WOLFXLSX.DataManager.get",
        lambda: dummy_data_manager,
    )

    WOLFXLSX(config).write_to_path(
        [
            Item.from_dict(
                {
                    "src": "原文",
                    "dst": "译文",
                    "row": 2,
                    "file_type": Item.FileType.WOLFXLSX,
                    "file_path": "wolf/not-sheet.xlsx",
                }
            )
        ]
    )

    output_file = (
        Path(dummy_data_manager.get_translated_path()) / "wolf" / "not-sheet.xlsx"
    )
    assert output_file.exists() is False
