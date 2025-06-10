import json
from enum import StrEnum

import openpyxl
import openpyxl.styles
import openpyxl.worksheet.worksheet
from PyQt5.QtCore import QModelIndex
from PyQt5.QtCore import Qt
from PyQt5.QtWidgets import QTableWidgetItem
from qfluentwidgets import TableWidget

class TableManager():

    class Type(StrEnum):

        GLOSSARY = "GLOSSARY"
        REPLACEMENT = "REPLACEMENT"
        TEXT_PRESERVE = "TEXT_PRESERVE"

    def __init__(self, type: str, data: list[dict[str, str]], table: TableWidget) -> None:
        super().__init__()

        # 初始化
        self.type = type
        self.data = data
        self.table = table

        # 更新中标识
        self.updating: bool = False

    # 重置
    def reset(self) -> None:
        self.data = []
        self.table.clearContents()
        self.table.horizontalHeader().setSortIndicator(-1, Qt.SortOrder.AscendingOrder)

    # 同步
    def sync(self) -> None:
        # 更新开始
        self.set_updating(True)

        # 去重
        dels: set[int] = set()
        for i in range(len(self.data)):
            for k in range(i + 1, len(self.data)):
                x = self.data[i]
                y = self.data[k]
                if x.get("src") == y.get("src"):
                    if x.get("dst") != "" and y.get("dst") == "":
                        dels.add(k)
                    elif x.get("dst") == "" and y.get("dst") == "" and x.get("info") != "" and y.get("info") == "":
                        dels.add(k)
                    elif x.get("dst") == "" and y.get("dst") == "" and x.get("regex") != "" and y.get("regex") == "":
                        dels.add(k)
                    else:
                        dels.add(i)
        self.data = [v for i, v in enumerate(self.data) if i not in dels]

        # 填充表格
        self.table.setRowCount(max(20, len(self.data) + 8))
        for row in range(self.table.rowCount()):
            for col in range(self.table.columnCount()):
                item = self.table.item(row, col)
                if item is not None:
                    item.setText("")
                else:
                    self.table.setItem(row, col, self.generate_item(col))

        # 遍历表格
        if self.type == __class__.Type.GLOSSARY:
            for row, v in enumerate(self.data):
                for col in range(self.table.columnCount()):
                    if col == 0:
                        self.table.item(row, col).setText(v.get("src", ""))
                    elif col == 1:
                        self.table.item(row, col).setText(v.get("dst", ""))
                    elif col == 2:
                        self.table.item(row, col).setText(v.get("info", ""))
        elif self.type == __class__.Type.REPLACEMENT:
            for row, v in enumerate(self.data):
                for col in range(self.table.columnCount()):
                    if col == 0:
                        self.table.item(row, col).setText(v.get("src", ""))
                    elif col == 1:
                        self.table.item(row, col).setText(v.get("dst", ""))
                    elif col == 2:
                        if v.get("regex", False) == True:
                            self.table.item(row, col).setText("✅")
                        else:
                            self.table.item(row, col).setText("")
        elif self.type == __class__.Type.TEXT_PRESERVE:
            for row, v in enumerate(self.data):
                for col in range(self.table.columnCount()):
                    if col == 0:
                        self.table.item(row, col).setText(v.get("src", ""))
                    elif col == 1:
                        self.table.item(row, col).setText(v.get("info", ""))

        # 更新结束
        self.set_updating(False)

    # 导出
    def export(self, path: str) -> None:
        # 新建工作表
        book: openpyxl.Workbook = openpyxl.Workbook()
        sheet: openpyxl.worksheet.worksheet.Worksheet = book.active

        # 设置表头
        sheet.column_dimensions["A"].width = 32
        sheet.column_dimensions["B"].width = 32
        sheet.column_dimensions["C"].width = 32
        sheet.column_dimensions["D"].width = 32
        TableManager.set_cell_value(sheet, 1, 1, "src", 10)
        TableManager.set_cell_value(sheet, 1, 2, "dst", 10)
        TableManager.set_cell_value(sheet, 1, 3, "info", 10)
        TableManager.set_cell_value(sheet, 1, 4, "regex", 10)

        # 将数据写入工作表
        for row, item in enumerate(self.data):
            TableManager.set_cell_value(sheet, row + 2, 1, item.get("src", ""), 10)
            TableManager.set_cell_value(sheet, row + 2, 2, item.get("dst", ""), 10)
            TableManager.set_cell_value(sheet, row + 2, 3, item.get("info", ""), 10)
            TableManager.set_cell_value(sheet, row + 2, 4, item.get("regex", ""), 10)

        # 保存工作簿
        book.save(f"{path}.xlsx")

        # 保存为 JSON
        with open(f"{path}.json", "w", encoding = "utf-8") as writer:
            writer.write(json.dumps(self.data, indent = 4, ensure_ascii = False))

    # 搜索
    def search(self, keyword: str, start: int) -> int:
        result: int = -1
        keyword = keyword.lower()

        # 从指定位置开始搜索
        for i, entry in enumerate(self.data):
            if i <= start:
                continue
            if any(keyword in v.lower() for v in entry.values() if isinstance(v, str)):
                result = i
                break

        # 如果未找到则从头开始搜索
        if result == -1:
            for i, entry in enumerate(self.data):
                if i > start:
                    continue
                if any(keyword in v.lower() for v in entry.values() if isinstance(v, str)):
                    result = i
                    break

        return result

    # 获取数据
    def get_data(self) -> list[dict[str, str]]:
        return self.data

    # 设置数据
    def set_data(self, data: list[dict[str, str]]) -> None:
        self.data = data

    # 获取更新中标识
    def get_updating(self) -> bool:
        return self.updating

    # 设置更新中标识
    def set_updating(self, updating: bool) -> None:
        self.updating = updating

    # 生成新的条目
    def generate_item(self, col: int) -> QTableWidgetItem:
        item = QTableWidgetItem("")
        item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)

        if self.type == __class__.Type.GLOSSARY:
            pass
        elif self.type == __class__.Type.REPLACEMENT:
            if col == 2:
                item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)
        elif self.type == __class__.Type.TEXT_PRESERVE:
            pass

        return item

    # 删除行事件
    def delete_row(self) -> None:
        selected_index = self.table.selectedIndexes()

        # 有效性检验
        if selected_index == None or len(selected_index) == 0:
            return

        # 逆序删除并去重以避免索引错误
        for row in sorted({item.row() for item in selected_index}, reverse = True):
            self.table.removeRow(row)

        # 删除行不会触发 itemChanged 事件，所以手动触发一下
        self.table.itemChanged.emit(QTableWidgetItem())

    # 切换正则事件
    def switch_regex(self) -> None:
        selected_index: list[QModelIndex] = self.table.selectedIndexes()

        # 有效性检验
        if selected_index == None or len(selected_index) == 0:
            return

        # 切换正则模式
        for row in {index.row() for index in selected_index}:
            item = self.table.item(row, 2)
            if item is None:
                item = QTableWidgetItem()
                self.table.setItem(row, 2, item)
            if item.text().strip() != "✅":
                item.setText("✅")
            else:
                item.setText("")

    # 获取行数据
    def get_entry_by_row(self, row: int) -> dict[str, str | bool]:
        items: list[QTableWidgetItem] = [
            self.table.item(row, col)
            for col in range(self.table.columnCount())
        ]

        if self.type == __class__.Type.GLOSSARY:
            return {
                "src": items[0].text().strip() if isinstance(items[0], QTableWidgetItem) else "",
                "dst": items[1].text().strip() if isinstance(items[1], QTableWidgetItem) else "",
                "info": items[2].text().strip() if isinstance(items[2], QTableWidgetItem) else "",
            }
        elif self.type == __class__.Type.REPLACEMENT:
            return {
                "src": items[0].text().strip() if isinstance(items[0], QTableWidgetItem) else "",
                "dst": items[1].text().strip() if isinstance(items[1], QTableWidgetItem) else "",
                "regex": items[2].text().strip() == "✅" if isinstance(items[2], QTableWidgetItem) else False,
            }
        elif self.type == __class__.Type.TEXT_PRESERVE:
            return {
                "src": items[0].text().strip() if isinstance(items[0], QTableWidgetItem) else "",
                "info": items[1].text().strip() if isinstance(items[1], QTableWidgetItem) else "",
            }

    # 从表格加载数据
    def append_data_from_table(self) -> None:
        for row in range(self.table.rowCount()):
            entry: dict[str, str | bool] = self.get_entry_by_row(row)
            if entry.get("src") != "":
                self.data.append(entry)

    # 从文件加载数据
    def append_data_from_file(self, path: str) -> None:
        result: list[dict[str, str]] = []

        if path.lower().endswith(".json"):
            result = self.load_from_json_file(path)
        elif path.lower().endswith(".xlsx"):
            result = self.load_from_xlsx_file(path)

        # 合并数据并去重
        self.data.extend(result)
        self.data = list({v["src"]: v for v in self.data}.values())

    # 从 json 文件加载数据
    def load_from_json_file(self, path: str) -> list[dict[str, str]]:
            result: list[dict[str, str]] = []

            # 读取文件
            inputs = []
            with open(path, "r", encoding = "utf-8-sig") as reader:
                inputs: dict[str, str] | list[dict[str, str]] = json.load(reader)

            # 标准字典列表
            # [
            #     {
            #         "key": "value",
            #         "key": "value",
            #         "key": "value",
            #     }
            # ]
            if isinstance(inputs, list):
                for entry in inputs:
                    # 格式校验
                    if isinstance(entry, dict) == False:
                        continue
                    if "src" not in entry:
                        continue

                    src: str = entry.get("src", "").strip()
                    if src != "":
                        result.append(
                            {
                                "src": src,
                                "dst": entry.get("dst", "").strip(),
                                "info": entry.get("info", "").strip(),
                                "regex": entry.get("regex", False),
                            }
                        )

            # Actors.json
            # [
            #     null,
            #     {
            #         "id": 1,
            #         "name": "レナリス",
            #         "nickname": "ローズ娼館の娼婦",
            #     },
            # ]
            if isinstance(inputs, list):
                for entry in inputs:
                    # 格式校验
                    if isinstance(entry, dict) == False:
                        continue
                    if isinstance(entry.get("id"), int) == False:
                        continue

                    id: int = entry.get("id", -1)
                    name: str = entry.get("name", "").strip()
                    nickname: str = entry.get("nickname", "").strip()

                    # 添加数据
                    if name != "":
                        result.append(
                            {
                                "src": f"\\n[{id}]",
                                "dst": name,
                                "info": "",
                                "regex": False,
                            }
                        )
                        result.append(
                            {
                                "src": f"\\N[{id}]",
                                "dst": name,
                                "info": "",
                                "regex": False,
                            }
                        )
                    if nickname != "":
                        result.append(
                            {
                                "src": f"\\nn[{id}]",
                                "dst": name,
                                "info": "",
                                "regex": False,
                            }
                        )
                        result.append(
                            {
                                "src": f"\\NN[{id}]",
                                "dst": name,
                                "info": "",
                                "regex": False,
                            }
                        )

            # 标准 KV 字典
            # {
            #     "ダリヤ": "达莉雅"
            # }
            if isinstance(inputs, dict):
                for k, v in inputs.items():
                    # 格式校验
                    if not isinstance(k, str):
                        continue

                    src: str = k.strip()
                    dst: str = v.strip() if v is not None else ""
                    if src != "":
                        result.append(
                            {
                                "src": src,
                                "dst": dst,
                                "info": "",
                                "regex": False,
                            }
                        )

            return result

    # 从 xlsx 文件加载数据
    def load_from_xlsx_file(self, path: str) -> list[dict]:
        result: list[dict[str, str]] = []

        sheet = openpyxl.load_workbook(path).active
        for row in range(1, sheet.max_row + 1):
            # 读取每一行的数据
            data: list[str] = [
                sheet.cell(row = row, column = col).value
                for col in range(1, 5)
            ]

            # 格式校验
            if not isinstance(data[0], str):
                continue

            src: str = data[0].strip()
            dst: str = data[1].strip() if data[1] is not None else ""
            info: str = data[2].strip() if data[2] is not None else ""
            regex: str = data[3].strip().lower() == "true" if data[3] is not None else False

            if src == "src" and dst == "dst":
                continue

            # 添加数据
            if src != "":
                result.append(
                    {
                        "src": src,
                        "dst": dst,
                        "info": info,
                        "regex": regex,
                    }
                )

        return result

    # 设置单元格值
    @classmethod
    def set_cell_value(cls, sheet: openpyxl.worksheet.worksheet.Worksheet, row: int, column: int, value: str, font_size: int = 9) -> None:
        if value is None:
            value = ""
        elif isinstance(value, str) == False:
            value = str(value)
        # 如果单元格内容以单引号 ' 开头，Excel 会将其视为普通文本而不是公式
        elif value.startswith("=") == True:
            value = "'" + value

        sheet.cell(row = row, column = column).value = value
        sheet.cell(row = row, column = column).font = openpyxl.styles.Font(size = font_size)
        sheet.cell(row = row, column = column).alignment  = openpyxl.styles.Alignment(wrap_text = True, vertical = "center", horizontal = "left")