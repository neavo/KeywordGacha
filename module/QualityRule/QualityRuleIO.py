from pathlib import Path
from typing import Any

import openpyxl
import openpyxl.worksheet.worksheet

from module.Utils.JSONTool import JSONTool
from module.Utils.SpreadsheetTool import SpreadsheetTool


class QualityRuleIO:
    """质量规则导入/导出。

    约束：
    - 纯逻辑层：不依赖 Qt
    - 规则存储形态保持与历史 JSON/XLSX 兼容（字段名不做强约束）
    """

    @staticmethod
    def load_rules_from_file(path: str) -> list[dict[str, Any]]:
        if not isinstance(path, str) or not path:
            return []

        lower = path.lower()
        if lower.endswith(".json"):
            return QualityRuleIO.load_from_json_file(path)
        if lower.endswith(".xlsx"):
            return QualityRuleIO.load_from_xlsx_file(path)
        return []

    @staticmethod
    def load_from_json_file(path: str) -> list[dict[str, Any]]:
        result: list[dict[str, Any]] = []

        inputs: Any = JSONTool.load_file(path)

        # 标准字典列表：[{"src": "...", "dst": "...", ...}, ...]
        if isinstance(inputs, list):
            for entry in inputs:
                if not isinstance(entry, dict):
                    continue
                if "src" not in entry:
                    continue

                src = str(entry.get("src", "")).strip()
                if not src:
                    continue

                result.append(
                    {
                        "src": src,
                        "dst": str(entry.get("dst", "")).strip(),
                        "info": str(entry.get("info", "")).strip(),
                        "regex": bool(entry.get("regex", False)),
                        "case_sensitive": bool(entry.get("case_sensitive", False)),
                    }
                )

        # Actors.json（RPG Maker）：[{"id": 1, "name": "...", "nickname": "..."}, ...]
        if isinstance(inputs, list):
            for entry in inputs:
                if not isinstance(entry, dict):
                    continue
                if not isinstance(entry.get("id"), int):
                    continue

                actor_id = int(entry.get("id", -1))
                name = str(entry.get("name", "")).strip()
                nickname = str(entry.get("nickname", "")).strip()

                if name:
                    for code in ("\\n", "\\N"):
                        result.append(
                            {
                                "src": f"{code}[{actor_id}]",
                                "dst": name,
                                "info": "",
                                "regex": False,
                                "case_sensitive": False,
                            }
                        )
                if nickname:
                    for code in ("\\nn", "\\NN"):
                        result.append(
                            {
                                "src": f"{code}[{actor_id}]",
                                "dst": nickname,
                                "info": "",
                                "regex": False,
                                "case_sensitive": False,
                            }
                        )

        # 标准 KV 字典：{"src": "dst", ...}
        if isinstance(inputs, dict):
            for k, v in inputs.items():
                if not isinstance(k, str):
                    continue
                src = k.strip()
                dst = str(v).strip() if v is not None else ""
                if not src:
                    continue
                result.append(
                    {
                        "src": src,
                        "dst": dst,
                        "info": "",
                        "regex": False,
                        "case_sensitive": False,
                    }
                )

        return result

    @staticmethod
    def load_from_xlsx_file(path: str) -> list[dict[str, Any]]:
        result: list[dict[str, Any]] = []

        sheet = openpyxl.load_workbook(path).active
        if not isinstance(sheet, openpyxl.worksheet.worksheet.Worksheet):
            return result

        for row in range(1, sheet.max_row + 1):
            # 读取每一行的 5 列（与历史 TableManager 导出列一致）
            data = [
                SpreadsheetTool.get_cell_value(sheet, row=row, column=col)
                for col in range(1, 6)
            ]

            if not data or data[0] is None:
                continue

            src = str(data[0]).strip()
            dst = str(data[1]).strip() if len(data) > 1 else ""
            info = str(data[2]).strip() if len(data) > 2 else ""
            regex = str(data[3]).strip().lower() == "true" if len(data) > 3 else False
            case_sensitive = (
                str(data[4]).strip().lower() == "true" if len(data) > 4 else False
            )

            # 跳过表头
            if src == "src" and dst == "dst":
                continue

            if not src:
                continue

            result.append(
                {
                    "src": src,
                    "dst": dst,
                    "info": info,
                    "regex": regex,
                    "case_sensitive": case_sensitive,
                }
            )

        return result

    @staticmethod
    def export_rules(path_base: str, rules: list[dict[str, Any]]) -> None:
        base = str(Path(path_base).with_suffix(""))

        # XLSX
        book: openpyxl.Workbook = openpyxl.Workbook()
        sheet = book.active
        if not isinstance(sheet, openpyxl.worksheet.worksheet.Worksheet):
            raise RuntimeError("Failed to create worksheet")

        sheet.column_dimensions["A"].width = 24
        sheet.column_dimensions["B"].width = 24
        sheet.column_dimensions["C"].width = 24
        sheet.column_dimensions["D"].width = 24
        sheet.column_dimensions["E"].width = 24

        SpreadsheetTool.set_cell_value(
            sheet, row=1, column=1, value="src", font_size=10
        )
        SpreadsheetTool.set_cell_value(
            sheet, row=1, column=2, value="dst", font_size=10
        )
        SpreadsheetTool.set_cell_value(
            sheet, row=1, column=3, value="info", font_size=10
        )
        SpreadsheetTool.set_cell_value(
            sheet, row=1, column=4, value="regex", font_size=10
        )
        SpreadsheetTool.set_cell_value(
            sheet, row=1, column=5, value="case_sensitive", font_size=10
        )

        for idx, item in enumerate(rules):
            row = idx + 2
            SpreadsheetTool.set_cell_value(
                sheet, row=row, column=1, value=item.get("src", ""), font_size=10
            )
            SpreadsheetTool.set_cell_value(
                sheet, row=row, column=2, value=item.get("dst", ""), font_size=10
            )
            SpreadsheetTool.set_cell_value(
                sheet, row=row, column=3, value=item.get("info", ""), font_size=10
            )
            SpreadsheetTool.set_cell_value(
                sheet, row=row, column=4, value=item.get("regex", ""), font_size=10
            )
            SpreadsheetTool.set_cell_value(
                sheet,
                row=row,
                column=5,
                value=item.get("case_sensitive", ""),
                font_size=10,
            )

        book.save(f"{base}.xlsx")

        # JSON
        JSONTool.save_file(f"{base}.json", rules, indent=4)
