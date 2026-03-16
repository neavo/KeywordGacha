import io
import os

import openpyxl
import openpyxl.worksheet.worksheet

from base.Base import Base
from model.Item import Item
from module.Config import Config
from module.Data.DataManager import DataManager
from module.Utils.SpreadsheetTool import SpreadsheetTool


class XLSX(Base):
    def __init__(self, config: Config) -> None:
        super().__init__()

        # 初始化
        self.config = config

    # 读取
    def read_from_path(self, abs_paths: list[str], input_path: str) -> list[Item]:
        items: list[Item] = []
        for abs_path in abs_paths:
            # 获取相对路径
            rel_path = os.path.relpath(abs_path, input_path)

            # 数据处理
            with open(abs_path, "rb") as reader:
                items.extend(self.read_from_stream(reader.read(), rel_path))

        return items

    # 从流读取
    def read_from_stream(self, content: bytes, rel_path: str) -> list[Item]:
        items: list[Item] = []

        # 数据处理
        book: openpyxl.Workbook = openpyxl.load_workbook(io.BytesIO(content))
        sheet = book.active

        if not isinstance(sheet, openpyxl.worksheet.worksheet.Worksheet):
            return items

        # 跳过空表格
        if sheet.max_row == 0 or sheet.max_column == 0:
            return items

        # 判断是否为 WOLF 翻译表格文件
        if self.is_wold_xlsx(sheet):
            return items

        for row in range(1, sheet.max_row + 1):
            src_val = sheet.cell(row=row, column=1).value
            dst_val = sheet.cell(row=row, column=2).value

            # 跳过读取失败的行
            # 数据不存在时为 None，存在时可能是 str int float 等多种类型
            if src_val is None:
                continue

            src = str(src_val)
            dst = str(dst_val) if dst_val is not None else ""

            if src == "":
                items.append(
                    Item.from_dict(
                        {
                            "src": src,
                            "dst": dst,
                            "row": row,
                            "file_type": Item.FileType.XLSX,
                            "file_path": rel_path,
                            "status": Base.ProjectStatus.EXCLUDED,
                        }
                    )
                )
            elif dst != "" and src != dst:
                items.append(
                    Item.from_dict(
                        {
                            "src": src,
                            "dst": dst,
                            "row": row,
                            "file_type": Item.FileType.XLSX,
                            "file_path": rel_path,
                            "status": Base.ProjectStatus.PROCESSED_IN_PAST,
                        }
                    )
                )
            else:
                items.append(
                    Item.from_dict(
                        {
                            "src": src,
                            "dst": dst,
                            "row": row,
                            "file_type": Item.FileType.XLSX,
                            "file_path": rel_path,
                            "status": Base.ProjectStatus.NONE,
                        }
                    )
                )

        return items

    # 写入
    def write_to_path(self, items: list[Item]) -> None:
        # 获取输出目录
        output_path = DataManager.get().get_translated_path()

        target = [item for item in items if item.get_file_type() == Item.FileType.XLSX]

        group: dict[str, list[Item]] = {}
        for item in target:
            group.setdefault(item.get_file_path(), []).append(item)

        # 分别处理每个文件
        for rel_path, group_items in group.items():
            # 按行号排序
            sorted_items = sorted(group_items, key=lambda x: x.get_row())

            # 新建工作表
            book: openpyxl.Workbook = openpyxl.Workbook()
            sheet = book.active

            if not isinstance(sheet, openpyxl.worksheet.worksheet.Worksheet):
                continue

            # 设置表头
            sheet.column_dimensions["A"].width = 64
            sheet.column_dimensions["B"].width = 64

            # 将数据写入工作表
            for item in sorted_items:
                row: int = item.get_row()
                SpreadsheetTool.set_cell_value(
                    sheet, row=row, column=1, value=item.get_src()
                )
                SpreadsheetTool.set_cell_value(
                    sheet, row=row, column=2, value=item.get_dst()
                )

            # 保存工作簿
            abs_path = os.path.join(output_path, rel_path)
            os.makedirs(os.path.dirname(abs_path), exist_ok=True)
            book.save(abs_path)

    # 是否为 WOLF 翻译表格文件
    def is_wold_xlsx(self, sheet: openpyxl.worksheet.worksheet.Worksheet) -> bool:
        value = sheet.cell(row=1, column=1).value
        if not isinstance(value, str) or "code" not in value.lower():
            return False

        value = sheet.cell(row=1, column=2).value
        if not isinstance(value, str) or "flag" not in value.lower():
            return False

        value = sheet.cell(row=1, column=3).value
        if not isinstance(value, str) or "type" not in value.lower():
            return False

        value = sheet.cell(row=1, column=4).value
        if not isinstance(value, str) or "info" not in value.lower():
            return False

        return True
