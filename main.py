import os
import re
import csv
import copy
import json
import asyncio
from types import SimpleNamespace

import jaconv
import openpyxl
import unicodedata

from rich import box
from rich.table import Table
from rich.prompt import Prompt
from rich.traceback import install

from model.LLM import LLM
from model.NER import NER
from model.Word import Word
from helper.LogHelper import LogHelper
from helper.TestHelper import TestHelper
from helper.TextHelper import TextHelper

# 定义常量
SCORE_THRESHOLD = 0.85

# 用于英文的代码段规则
CODE_PATTERN_EN = (
    "" + r"if\(.{0,10}[vs]\[\d+\].{0,10}\)" + "",           # if(!s[982]) if(v[982] >= 1)
    "" + r"en\(.{0,10}[vs]\[\d+\].{0,10}\)" + "",           # en(!s[982]) en(v[982] >= 1)
    "" + r"[/\\][a-z]{1,5}<[\d]{0,10}>" + "",               # /C<1> \FS<12>
    "" + r"[/\\][a-z]{1,5}\[[\d]{0,10}\]" + "",             # /C[1] \FS[12]
    "" + r"[/\\][a-z]{1,5}(?=<.{0,10}>)" + "",              # /C<非数字> /C<非数字> \FS<非数字> \FS<非数字> 中的前半部分
    "" + r"[/\\][a-z]{1,5}(?=\[.{0,10}\])" + "",            # /C[非数字] /C[非数字] \FS[非数字] \FS[非数字] 中的前半部分
)

# 用于非英文的代码段规则
CODE_PATTERN_NON_EN = (
    "" + r"if\(.{0,10}[vs]\[\d+\].{0,10}\)" + "",           # if(!s[982]) if(v[982] >= 1)
    "" + r"en\(.{0,10}[vs]\[\d+\].{0,10}\)" + "",           # en(!s[982]) en(v[982] >= 1)
    "" + r"[/\\][a-z]{1,5}<[\da-z]{0,10}>" + "",            # /C<y> /C<1> \FS<xy> \FS<12>
    "" + r"[/\\][a-z]{1,5}\[[\da-z]{0,10}\]" + "",          # /C[x] /C[1] \FS[xy] \FS[12]
    "" + r"[/\\][a-z]{1,5}(?=<.{0,10}>)" + "",              # /C<非数字非字母> /C<非数字非字母> \FS<非数字非字母> \FS<非数字非字母> 中的前半部分
    "" + r"[/\\][a-z]{1,5}(?=\[.{0,10}\])" + "",            # /C[非数字非字母] /C[非数字非字母] \FS[非数字非字母] \FS[非数字非字母] 中的前半部分
)

# 同时作用于英文于非英文的代码段规则
CODE_PATTERN_COMMON = (
    "" + r"\\fr" + "",                                      # 重置文本的改变
    "" + r"\\fb" + "",                                      # 加粗
    "" + r"\\fi" + "",                                      # 倾斜
    "" + r"\\\{" + "",                                      # 放大字体 \{
    "" + r"\\\}" + "",                                      # 缩小字体 \}
    "" + r"\\g" + "",                                       # 显示货币 \G
    "" + r"\\\$" + "",                                      # 打开金币框 \$
    "" + r"\\\." + "",                                      # 等待0.25秒 \.
    "" + r"\\\|" + "",                                      # 等待1秒 \|
    "" + r"\\!" + "",                                       # 等待按钮按下 \!
    # "" + r"\\>" + "",                                     # 在同一行显示文字 \>
    # "" + r"\\<" + "",                                     # 取消显示所有文字 \<
    "" + r"\\\^" + "",                                      # 显示文本后不需要等待 \^
    # "" + r"\\n" + "",                                     # 换行符 \\n
    "" + r"\r\n" + "",                                      # 换行符 \r\n
    "" + r"\n" + "",                                        # 换行符 \n
    "" + r"\\\\<br>" + "",                                  # 换行符 \\<br>
    "" + r"<br>" + "",                                      # 换行符 <br>
)

# 加载 Actors.json 数据
def load_names(path: str) -> tuple[dict, dict]:
    names = {}
    nicknames = {}

    if not os.path.exists(path):
        pass
    else:
        with open(path, "r", encoding = "utf-8") as reader:
            for item in json.load(reader):
                if isinstance(item, dict):
                    id = item.get("id", -1)

                    if not isinstance(id, int):
                        continue

                    names[id] = item.get("name", "")
                    nicknames[id] = item.get("nickname", "")
        LogHelper.info(f"从 [green]Actors.json[/] 文件中加载了 {len(names) + len(nicknames)} 条数据，稍后将执行 [green]角色代码还原[/] 步骤 ...")

    return names, nicknames

# 执行替换
def do_replace(match: re.Match, names: dict) -> str:
    i = int(match.group(1))

    # 索引在范围内则替换，不在范围内则原文返回
    if i in names:
        return names.get(i, "")
    else:
        return match.group(0)

# 还原角色代码
def replace_name_code(text: str, names: dict, nicknames: dict) -> str:
    # 根据 actors 中的数据还原 角色代码 \N[123] 实际指向的名字
    text = re.sub(
        r"\\n\[(\d+)\]",
        lambda match: do_replace(match, names),
        text,
        flags = re.IGNORECASE
    )

    # 根据 actors 中的数据还原 角色代码 \NN[123] 实际指向的名字
    text = re.sub(
        r"\\nn\[(\d+)\]",
        lambda match: do_replace(match, nicknames),
        text,
        flags = re.IGNORECASE
    )

    return text

# 清理文本
def cleanup(line: str, language: int, names: dict, nicknames: dict) -> str:
    # 还原角色代码
    line = replace_name_code(line, names, nicknames)

    # 将 队伍成员代码 \P[123] 替换为 teammate_123
    line = re.sub(r"\\[P]\[(\d+)\]", r"teammate_\1", line, flags = re.IGNORECASE)

    if language == NER.LANGUAGE.EN:
        line = re.sub(rf"(?:{"|".join(CODE_PATTERN_EN + CODE_PATTERN_COMMON)})+", "", line, flags = re.IGNORECASE)
    else:
        line = re.sub(rf"(?:{"|".join(CODE_PATTERN_NON_EN + CODE_PATTERN_COMMON)})+", "", line, flags = re.IGNORECASE)

    # 由于上面的代码移除，可能会产生空人名框的情况，干掉
    line = line.replace("【】", "")

    # 干掉除了空格以外的行内空白符（包括换行符、制表符、回车符、换页符等）
    line = re.sub(r"[^\S ]+", "", line)

    # 合并连续的空格为一个空格
    line = re.sub(r" +", " ", line)

    return line

# 读取 .txt 文件
def read_txt_file(path: str) -> tuple[list, list]:
    lines = []
    names = []
    encodings = ["utf-8", "utf-16", "shift-jis"]

    for encoding in encodings:
        try:
            with open(path, "r", encoding = encoding) as file:
                return file.readlines(), []
        except UnicodeDecodeError as e:
            LogHelper.debug(f"使用 {encoding} 编码读取数据文件时发生错误 - {e}")
        except Exception as e:
            LogHelper.error(f"读取数据文件时发生错误 - {LogHelper.get_trackback(e)}")
            break

    return lines, names

# 读取 .csv 文件
def read_csv_file(path: str) -> tuple[list, list]:
    lines = []
    names = []

    try:
        with open(path, "r", newline = "", encoding = "utf-8") as file:
            for row in csv.reader(file):
                lines.append(row[0])
    except Exception as e:
        LogHelper.error(f"读取数据文件时发生错误 - {LogHelper.get_trackback(e)}")

    return lines, names

# 读取 .json 文件
def read_json_file(path: str) -> tuple[list, list]:
    lines = []
    names = []

    try:
        # 读取并加载JSON文件
        with open(path, "r", encoding = "utf-8") as file:
            datas = json.load(file)

            # 针对 MTool 导出文本
            # {
            #   "「お前かよ。開けて。着替えなきゃ」" : "「お前かよ。開けて。着替えなきゃ」"
            # }
            if isinstance(datas, dict):
                for k, v in datas.items():
                    if isinstance(k, str) and isinstance(v, str):
                        lines.append(k)

            # 针对 SExtractor 导出的带name字段JSON数据
            # [{
            #   "name": "少年",
            #   "message": "「お前かよ。開けて。着替えなきゃ」"
            # }]
            if isinstance(datas, list):
                # 确保数据是字典类型
                for data in [data for data in datas if isinstance(data, dict)]:
                    name = data.get("name", "").strip()
                    message = data.get("message", "").strip()

                    if message == "":
                        continue

                    if name == "":
                        lines.append(f"{message}")
                    else:
                        message = f"【{name}】{message}"
                        names.append((name, message))
                        lines.append(message)
    except Exception as e:
        LogHelper.error(f"读取数据文件时发生错误 - {LogHelper.get_trackback(e)}")

    return lines, names

# 读取 .xlsx 文件
def read_xlsx_file(path: str) -> tuple[list, list]:
    lines = []
    names = []

    try:
        sheet = openpyxl.load_workbook(path).active
        for row in range(1, sheet.max_row + 1):
            cell_01 = sheet.cell(row = row, column = 1).value
            if isinstance(cell_01, str) and cell_01 != "":
                lines.append(cell_01)
    except Exception as e:
        LogHelper.error(f"读取数据文件时发生错误 - {LogHelper.get_trackback(e)}")

    return lines, names

# 读取文件
def read_file(path: str) -> tuple[list, list]:
    lines_ex = []
    names_ex = []

    if path.endswith(".txt"):
        lines_ex, names_ex = read_txt_file(path)
    elif path.endswith(".csv"):
        lines_ex, names_ex = read_csv_file(path)
    elif path.endswith(".json"):
        lines_ex, names_ex = read_json_file(path)
    elif path.endswith(".xlsx"):
        lines_ex, names_ex = read_xlsx_file(path)

    return lines_ex, names_ex

# 从输入文件中加载数据
def load_lines_from_input_file(language: int) -> tuple[list, list, str]:
    # 从 input 目录内寻找目标文件
    paths = []
    if os.path.isdir("input"):
        paths = [
            entry.path
            for entry in os.scandir("input")
            if entry.is_file() and entry.path.endswith((".txt", ".csv", ".json", ".xlsx"))
        ]

    # 分别处理找到和没找到的情况
    if len(paths) == 0:
        input_path = LogHelper.input("请输入数据文件的路径: ").strip('"')
    else:
        user_input = LogHelper.input(f"已在 [green]input[/] 路径下找到数据文件 [green]{len(paths)}[/] 个，按回车直接使用或输入其他路径：").strip('"')
        input_path = user_input if len(user_input) > 0 else "input"
    LogHelper.print("")

    # 分别处理输入路径是文件和文件夹的情况
    paths = []
    if os.path.isfile(input_path):
        paths = [input_path]
    elif os.path.isdir(input_path):
        paths = [entry.path for entry in os.scandir(input_path)]
    paths = [path for path in paths if path.endswith((".txt", ".csv", ".json", ".xlsx"))]

    # 尝试从输入路径的同级路径或者下级路径加载角色数据
    names, nicknames = {}, {}
    if os.path.isfile(f"{input_path}/Actors.json"):
        names, nicknames = load_names(f"{input_path}/Actors.json")
    elif os.path.isfile(f"{os.path.dirname(input_path)}/Actors.json"):
        names, nicknames = load_names(f"{os.path.dirname(input_path)}/Actors.json")

    # 依次读取每个数据文件
    with LogHelper.status("正在读取输入文件 ..."):
        input_lines, input_names = [], []
        for path in paths:
            lines_ex, names_ex = read_file(path)
            input_lines.extend(lines_ex)
            input_names.extend(names_ex)

        input_names_filtered = []
        for name, message in input_names:
            input_names_filtered.append((name, cleanup(message, language, names, nicknames)))

        input_lines_filtered = []
        for line in input_lines:
            line = cleanup(line, language, names, nicknames)

            if len(line) == 0:
                continue

            if language == NER.LANGUAGE.EN:
                line = unicodedata.normalize("NFKC", line)

            if language == NER.LANGUAGE.JP:
                line = jaconv.normalize(line, mode = "NFKC")

            if language == NER.LANGUAGE.ZH and not TextHelper.has_any_cjk(line):
                continue

            if language == NER.LANGUAGE.EN and not TextHelper.has_any_latin(line):
                continue

            if language == NER.LANGUAGE.JP and not TextHelper.has_any_japanese(line):
                continue

            if language == NER.LANGUAGE.KO and not TextHelper.has_any_korean(line):
                continue

            input_lines_filtered.append(line.strip())

    LogHelper.info(f"已读取到文本 {len(input_lines)} 行，其中有效文本 {len(input_lines_filtered)} 行, 角色名 {len(input_names_filtered)} 个...")
    return input_lines_filtered, input_names_filtered, input_path

# 合并词语，并按出现次数排序
def merge_words(words: list[Word]) -> list[Word]:
    words_unique = {}
    for word in words:
        key = (word.surface, word.ner_type) # 只有文字和类型都一样才视为相同条目，避免跨类词条目合并
        if key not in words_unique:
            words_unique[key] = []
        words_unique[key].append(word)

    words_merged = []
    for v in words_unique.values():
        word = v[0]
        word.context = list(set([word.context[0] for word in v if word.context[0] != ""]))
        word.context.sort(key = lambda x: len(x), reverse = True)
        word.count = len(word.context)
        word.score = min(0.9999, sum(w.score for w in v) / len(v))
        words_merged.append(word)

    return sorted(words_merged, key = lambda x: x.count, reverse = True)

# 按置信度过滤词语
def filter_words_by_score(words: list[Word], threshold: float) -> list[Word]:
    return [word for word in words if word.score >= threshold]

# 按出现次数过滤词语
def filter_words_by_count(words: list[Word], threshold: float) -> list[Word]:
    return [word for word in words if word.count >= max(1, threshold)]

# 获取指定类型的词
def get_words_by_ner_type(words: list[Word], ner_type: str) -> list[Word]:
    # 显式的复制对象，避免后续修改对原始列表的影响，浅拷贝不复制可变对象（列表、字典、自定义对象等），慎重修改它们
    return [copy.copy(word) for word in words if word.ner_type == ner_type]

# 移除指定类型的词
def remove_words_by_ner_type(words: list[Word], ner_type: str) -> list[Word]:
    return [word for word in words if word.ner_type != ner_type]

# 指定类型的词
def replace_words_by_ner_type(words: list[Word], in_words: list[Word], ner_type: str) -> list[Word]:
    words = remove_words_by_ner_type(words, ner_type)
    words.extend(in_words)
    return words

# 将 词语日志 写入文件
def write_words_log_to_file(words: list[Word], path, language: int) -> None:
    with open(path, "w", encoding = "utf-8") as file:
        for k, word in enumerate(words):
            if getattr(word, "surface", "") != "":
                file.write(f"词语原文 : {word.surface}\n")

            if getattr(word, "score", float(-1)) >= 0:
                file.write(f"置信度 : {word.score:.4f}\n")

            if getattr(word, "surface_romaji", "") != "":
                if language == NER.LANGUAGE.JP:
                    file.write(f"罗马音 : {word.surface_romaji}\n")

            if getattr(word, "count", int(-1)) >= 0:
                file.write(f"出现次数 : {word.count}\n")

            if len(getattr(word, "surface_translation", "")) > 0:
                if word.surface_translation_description == "":
                    file.write(f"词语翻译 : {word.surface_translation}\n")
                else:
                    file.write(f"词语翻译 : {word.surface_translation}, {word.surface_translation_description}\n")

            if getattr(word, "attribute", "") != "":
                file.write(f"角色性别 : {word.attribute}\n")

            if getattr(word, "context_summary", "") != "":
                file.write(f"语义分析 : {word.context_summary}\n")

            if len(getattr(word, "context", [])) > 0:
                file.write("上下文原文 : ※※※※※※※※※※※※※※※※※※※※※※※※※※※※※※※※※※※※※※※※※※※※※※※※※※※※※※※※※※\n")
                file.write(f"{word.get_context_str_for_translate(language)}\n")

            if len(getattr(word, "context_translation", [])) > 0:
                file.write("上下文翻译 : ※※※※※※※※※※※※※※※※※※※※※※※※※※※※※※※※※※※※※※※※※※※※※※※※※※※※※※※※※※\n")
                file.write(f"{"\n".join(word.context_translation)}\n")

            if LogHelper.is_debug():
                if word.llmresponse_summarize_context != "":
                    file.write(f"{word.llmresponse_summarize_context}\n")
                if word.llmresponse_translate_context != "":
                    file.write(f"{word.llmresponse_translate_context}\n")
                if word.llmresponse_translate_surface != "":
                    file.write(f"{word.llmresponse_translate_surface}\n")

            # 多写入一个换行符，确保每段信息之间有间隔
            file.write("\n")

    LogHelper.info(f"结果已写入 - [green]{path}[/]")

# 将 词语列表 写入文件
def write_words_list_to_file(words: list[Word], path, language: int) -> None:
    with open(path, "w", encoding = "utf-8") as file:
        data = {}
        for k, word in enumerate(words):
            data[word.surface] = word.surface_translation

        file.write(json.dumps(data, indent = 4, ensure_ascii = False))
        LogHelper.info(f"结果已写入 - [green]{path}[/]")

# 将 AiNiee 词典写入文件
def write_ainiee_dict_to_file(words: list[Word], path, language: int) -> None:
    type_map = {
        "PER": "角色",      # 表示人名，如"张三"、"约翰·多伊"等。
        "ORG": "组织",      # 表示组织，如"联合国"、"苹果公司"等。
        "LOC": "地点",      # 表示地点，通常指非地理政治实体的地点，如"房间"、"街道"等。
        "PRD": "物品",      # 表示产品，如"iPhone"、"Windows操作系统"等。
        "EVT": "事件",      # 表示事件，如"奥运会"、"地震"等。
    }

    with open(path, "w", encoding = "utf-8") as file:
        datas = []
        for word in words:
            if word.surface_translation == "":
                continue

            data = {}
            data["src"] = word.surface
            data["srt"] = word.surface
            data["dst"] = word.surface_translation

            if word.ner_type == "PER" and "男" in word.attribute:
                data["info"] = "男性的名字"
            elif word.ner_type == "PER" and "女" in word.attribute:
                data["info"] = "女性的名字"
            elif word.ner_type == "PER":
                data["info"] = "名字"
            else:
                data["info"] = f"{type_map.get(word.ner_type)}的名字"

            datas.append(data)

        file.write(json.dumps(datas, indent = 4, ensure_ascii = False))
        LogHelper.info(f"结果已写入 - [green]{path}[/]")

# 将 GalTransl 词典写入文件
def write_galtransl_dict_to_file(words: list[Word], path, language: int) -> None:
    type_map = {
        "PER": "角色",      # 表示人名，如"张三"、"约翰·多伊"等。
        "ORG": "组织",      # 表示组织，如"联合国"、"苹果公司"等。
        "LOC": "地点",      # 表示地点，通常指非地理政治实体的地点，如"房间"、"街道"等。
        "PRD": "物品",      # 表示产品，如"iPhone"、"Windows操作系统"等。
        "EVT": "事件",      # 表示事件，如"奥运会"、"地震"等。
    }

    with open(path, "w", encoding = "utf-8") as file:
        for word in words:
            if word.surface_translation == "":
                continue

            line = f"{word.surface}\t{word.surface_translation}"

            if word.ner_type == "PER" and "男" in word.attribute:
                line = line + "\t男性的名字"
            elif word.ner_type == "PER" and "女" in word.attribute:
                line = line + "\t女性的名字"
            elif word.ner_type == "PER":
                line = line + "\t名字"
            else:
                line = line + f"\t{type_map.get(word.ner_type)}的名字"

            file.write(f"{line}\n")
        LogHelper.info(f"结果已写入 - [green]{path}[/]")

# 开始处理文本
async def process_text(llm: LLM, ner: NER, config: SimpleNamespace, language: int) -> None:
    # 设置 LLM 语言
    llm.set_language(language)

    # 选择处理模式
    # 中文没有翻译的过程，所以无法支持快速模式
    if language == NER.LANGUAGE.ZH:
        process_mode = LLM.PROCESS_MODE.NORMAL
    else:
        process_mode = print_menu_process_mode()

    # 读取输入文件
    input_lines, input_names, config.input_path = load_lines_from_input_file(language)

    # 查找 NER 实体
    LogHelper.info("即将开始执行 [查找 NER 实体] ...")
    words = []
    words = ner.search_for_entity(input_lines, input_names, language)

    # 等待 还原词根任务 结果，只对日文启用
    if language == NER.LANGUAGE.JP:
        LogHelper.info("即将开始执行 [还原词根] ...")
        words = ner.lemmatize_words_by_morphology(words)
        words = remove_words_by_ner_type(words, "")
        LogHelper.info("[还原词根] 已完成 ...")

    # 合并相同词条
    words = merge_words(words)

    # 调试模式时，检查置信度阈值
    if LogHelper.is_debug():
        with LogHelper.status("正在检查置信度阈值 ..."):
            TestHelper.check_score_threshold(words, "check_score_threshold.log")

    # 按出现次数阈值进行筛选
    LogHelper.info(f"即将开始执行 [阈值过滤] ... 当前出现次数的阈值设置为 {config.count_threshold} ...")
    words = filter_words_by_score(words, SCORE_THRESHOLD)
    words = filter_words_by_count(words, config.count_threshold)
    LogHelper.info("[阈值过滤] 已完成 ...")

    # 设置请求限制器
    llm.set_request_limiter()

    # 等待翻译词语任务结果
    # 中文无需翻译
    if language != NER.LANGUAGE.ZH:
        LogHelper.info("即将开始执行 [词语翻译] ...")
        words = await llm.translate_surface_batch(words)
        words = remove_words_by_ner_type(words, "")

    # 调试模式时，前置检查结果重复度
    if LogHelper.is_debug():
        with LogHelper.status("正在检查结果重复度..."):
            TestHelper.check_result_duplication(
                [word for word in words if llm.check_keyword_in_description(word, None)],
                 "check_result_duplication_01.log",
            )

    # 等待 语义分析任务 结果
    LogHelper.info("即将开始执行 [语义分析] ...")
    words_person = get_words_by_ner_type(words, "PER")
    words_person = await llm.summarize_context_batch(words_person, process_mode)
    words = replace_words_by_ner_type(words, words_person, "PER")
    words = remove_words_by_ner_type(words, "")

    # 调试模式时，后置检查结果重复度
    if LogHelper.is_debug():
        with LogHelper.status("正在检查结果重复度..."):
            TestHelper.check_result_duplication(words, "check_result_duplication_02.log")

    ner_type = {
        "PER": "角色实体",
        "ORG": "组织实体",
        "LOC": "地点实体",
        "PRD": "物品实体",
        "EVT": "事件实体",
    }

    # 等待 上下文翻译 任务结果
    if language in (NER.LANGUAGE.EN, NER.LANGUAGE.JP, NER.LANGUAGE.KO):
        for k, v in ner_type.items():
            if (
                (k == "PER" and config.translate_context_per == 1)
                or (k != "PER" and config.translate_context_other == 1)
            ):
                LogHelper.info(f"即将开始执行 [上下文翻译 - {v}] ...")
                word_type = get_words_by_ner_type(words, k)
                word_type = await llm.translate_context_batch(word_type)
                words = replace_words_by_ner_type(words, word_type, k)

    # 将结果写入文件
    dir_name, file_name_with_extension = os.path.split(config.input_path)
    file_name, extension = os.path.splitext(file_name_with_extension)

    LogHelper.info("")
    os.makedirs("output", exist_ok = True)
    for k, v in ner_type.items():
        words_ner_type = get_words_by_ner_type(words, k)
        os.remove(f"output/{file_name}_{v}_日志.txt") if os.path.exists(f"output/{file_name}_{v}_日志.txt") else None
        os.remove(f"output/{file_name}_{v}_列表.json") if os.path.exists(f"output/{file_name}_{v}_列表.json") else None
        os.remove(f"output/{file_name}_{v}_ainiee.json") if os.path.exists(f"output/{file_name}_{v}_ainiee.json") else None
        os.remove(f"output/{file_name}_{v}_galtransl.txt") if os.path.exists(f"output/{file_name}_{v}_galtransl.txt") else None

        if len(words_ner_type) > 0:
            write_words_log_to_file(words_ner_type, f"output/{file_name}_{v}_日志.txt", language)
            write_words_list_to_file(words_ner_type, f"output/{file_name}_{v}_列表.json", language)
            write_ainiee_dict_to_file(words_ner_type, f"output/{file_name}_{v}_ainiee.json", language)
            write_galtransl_dict_to_file(words_ner_type, f"output/{file_name}_{v}_galtransl.txt", language)

    # 等待用户退出
    LogHelper.info("")
    LogHelper.info("工作流程已结束 ... 请检查生成的数据文件 ...")
    LogHelper.info("")
    LogHelper.info("")
    os.system("pause")

# 接口测试
async def test_api(llm: LLM) -> None:
    # 设置请求限制器
    llm.set_request_limiter()

    # 等待接口测试结果
    if await llm.api_test():
        LogHelper.print("")
        LogHelper.info("接口测试 [green]执行成功[/] ...")
    else:
        LogHelper.print("")
        LogHelper.warning("接口测试 [red]执行失败[/], 请检查配置文件 ...")

    LogHelper.print("")
    os.system("pause")
    os.system("cls")

# 打印应用信息
def print_app_info(config: SimpleNamespace) -> None:
    LogHelper.print()
    LogHelper.print()
    LogHelper.rule("KeywordGacha", style = "light_goldenrod2")
    LogHelper.rule("[blue]https://github.com/neavo/KeywordGacha", style = "light_goldenrod2")
    LogHelper.rule("使用 OpenAI 兼容接口自动生成小说、漫画、字幕、游戏脚本等任意文本中的词语表的翻译辅助工具", style = "light_goldenrod2")
    LogHelper.print()

    table = Table(
        box = box.ASCII2,
        expand = True,
        highlight = True,
        show_lines = True,
        border_style = "light_goldenrod2"
    )
    table.add_column("设置", style = "white", ratio = 1, overflow = "fold")
    table.add_column("当前值", style = "white", ratio = 1, overflow = "fold")
    table.add_column("设置", style = "white", ratio = 1, overflow = "fold")
    table.add_column("当前值", style = "white", ratio = 1, overflow = "fold")

    rows = [
        ("接口密钥", str(config.api_key), "模型名称", str(config.model_name)),
        ("接口地址", str(config.base_url)),
        ("是否翻译角色实体上下文", "是" if config.translate_context_per == 1 else "否", "是否翻译其他实体上下文", "是" if config.translate_context_other == 1 else "否"),
        ("网络请求超时时间", f"{config.request_timeout} 秒" , "网络请求频率阈值", f"{config.request_frequency_threshold} 次/秒"),
    ]

    for row in rows:
        table.add_row(*row)

    LogHelper.print(table)
    LogHelper.print()
    LogHelper.print("请编辑 [green]config.json[/] 文件来修改应用设置 ...")
    LogHelper.print()

# 打印菜单
def print_menu_main() -> int:
    LogHelper.print("请选择功能：")
    LogHelper.print("")
    LogHelper.print("\t--> 1. 开始处理 [green]中文文本[/]")
    LogHelper.print("\t--> 2. 开始处理 [green]英文文本[/]")
    LogHelper.print("\t--> 3. 开始处理 [green]日文文本[/]")
    LogHelper.print("\t--> 4. 开始处理 [green]韩文文本（初步支持）[/]")
    LogHelper.print("\t--> 5. 开始执行 [green]接口测试[/]")
    LogHelper.print("")
    choice = int(Prompt.ask("请输入选项前的 [green]数字序号[/] 来使用对应的功能，默认为 [green][3][/] ",
        choices = ["1", "2", "3", "4", "5"],
        default = "3",
        show_choices = False,
        show_default = False
    ))
    LogHelper.print("")

    return choice

# 打印处理模式菜单
def print_menu_process_mode() -> int:
    LogHelper.print("请选择处理模式：")
    LogHelper.print("")
    LogHelper.print("\t--> 1. [green]普通模式[/]：速度较慢，通过语义分析对结果进行确认，理论上可以提供最为精确的处理结果")
    LogHelper.print("\t--> 2. [green]快速模式[/]：跳过语义分析步骤，速度快，消耗 Token 少，结果中将不包含角色性别与故事总结信息")
    LogHelper.print("")
    choice = int(Prompt.ask("请输入选项前的 [green]数字序号[/] 来使用对应的处理模式，默认为 [green][1][/] ",
        choices = [f"{LLM.PROCESS_MODE.NORMAL}", f"{LLM.PROCESS_MODE.QUICK}"],
        default = f"{LLM.PROCESS_MODE.NORMAL}",
        show_choices = False,
        show_default = False
    ))
    LogHelper.print("")

    return choice

# 主函数
async def begin(llm: LLM, ner: NER, config: SimpleNamespace) -> None:
    choice = -1
    while choice not in [1, 2, 3, 4]:
        print_app_info(config)

        choice = print_menu_main()
        if choice == 1:
            await process_text(llm, ner, config, NER.LANGUAGE.ZH)
        elif choice == 2:
            await process_text(llm, ner, config, NER.LANGUAGE.EN)
        elif choice == 3:
            await process_text(llm, ner, config, NER.LANGUAGE.JP)
        elif choice == 4:
            await process_text(llm, ner, config, NER.LANGUAGE.KO)
        elif choice == 5:
            await test_api(llm)

# 一些初始化步骤
def load_config() -> tuple[LLM, NER, SimpleNamespace]:
    with LogHelper.status("正在初始化 [green]KG[/] 引擎 ..."):
        try:
            config = SimpleNamespace()
            config_file = "config_dev.json" if os.path.exists("config_dev.json") else "config.json"
            with open(config_file, "r", encoding = "utf-8") as file:
                for k, v in json.load(file).items():
                    setattr(config, k, v[0])
        except FileNotFoundError:
            LogHelper.error(f"文件 {config_file} 未找到.")
        except json.JSONDecodeError:
            LogHelper.error(f"文件 {config_file} 不是有效的JSON格式.")

        # 初始化 LLM 对象
        llm = LLM(config)
        llm.load_prompt()

        # 初始化 NER 对象
        ner = NER()
        ner.load_blacklist()

    return llm, ner, config

# 确保程序出错时可以捕捉到错误日志
async def main() -> None:
    try:
        # 注册全局异常追踪器
        install()

        # 加载配置
        llm, ner, config = load_config()

        # 开始处理
        await begin(llm, ner, config)
    except EOFError:
        LogHelper.error("EOFError - 程序即将退出 ...")
    except KeyboardInterrupt:
        LogHelper.error("KeyboardInterrupt - 程序即将退出 ...")
    except Exception as e:
        LogHelper.error(f"{LogHelper.get_trackback(e)}")
        LogHelper.print()
        LogHelper.print()
        LogHelper.error("出现严重错误，程序即将退出，错误信息已保存至日志文件 [green]KeywordGacha.log[/] ...")
        LogHelper.print()
        LogHelper.print()
        os.system("pause")

# 入口函数
if __name__ == "__main__":
    asyncio.run(main())